# encoding: utf-8
# @author: w4dll
# @file: mytool.py
# @time: 2020/4/19 19:20

import json
import re, base64
from openpyxl import load_workbook, workbook
import random

# 从文件中读取数据
def readjson(filepath):
    try:
        with open(filepath, "r") as file:
            data = json.load(file)
            file.close()
            return data
    except :
        return ""

# 将数据写入json文件中
def writejson(filepath, data):
    with open(filepath, "w") as file:
        json.dump(data, file)
        file.close()

# todo 将excel文件转换成题库字典文件并返回
def importExcelTk(file_path):

    a, b = re.split('[/ -.]', file_path)[-3:-1]  # 分离出单位、专业

    # 读入excel信息
    wb = load_workbook(file_path)
    sheet = wb.worksheets[0]

    # 将每行数据添加到题库字典中
    step = 0
    data_tiku = []              # 存储导入的题库信息
    for row in list(sheet.rows)[1:]:  # 去掉首行标题
        val, tmp = [], ""
        for cell in row:
            val.append(cell.value)

        # todo 这里可以对导入判断等尽心优化处理，容错，题型纠错等

        # 去空后存储每行数据
        s = [a, b]
        for i in range(len(val)-1):
            if val[i] != None and str(val[i]).strip() != "":
                s.append(val[i])
        # s.append(val[-2])
        if val[-1] == None:
            s.append("")
        else:
            s.append(val[-1])

        step = step + 1
        data_tiku.append(s)
    # print(data_tiku)
    return data_tiku

# 将题库文件中的题库导入到excel表格
def writeToExcel(str_filepath, dict_data):

    wb = workbook.Workbook()

    # todo 在题库中插入题目类型、专业、单位等信息，取最长列列数，将每行长度补齐
    length_row = 0
    sheet_list = []
    for x in dict_data:
        length_row = max(length_row, len(dict_data[x]))
        sheet_list.append(dict_data[x])

    # todo 将题库补全，长度统一
    for i in range(0, len(sheet_list)):
        if len(sheet_list[i]) < length_row:
            for j in range(0, length_row - len(sheet_list[i])):
                sheet_list[i].insert(len(sheet_list[i])-2, "")

    # todo 插入表头信息
    first_row = ["适用人员", "适用单位", "题目", "题型"]
    i=1
    for m in range(4, length_row-2):
        first_row.append("选项"+str(i))
        i +=1
    first_row.append("答案")
    first_row.append("备注")
    sheet_list.insert(0, first_row)

    # todo 将每行数据写入excel表格
    for x in sheet_list:
        wb.worksheets[0].append(x)
    try:
        wb.save(str_filepath)
        return True
    except FileNotFoundError:
        print("导出题库时，excel文件保存失败！")
        return False

# 返回一个不重复的随机索引序列
def get_random(num, num_max):
    list_num = []
    if num >= num_max:
        return range(num_max)
    while(len(list_num) < num):
        a = random.randint(0, num_max-1)
        if a not in list_num:
            list_num.append(a)
    return list_num


if __name__ == '__main__':

    a = get_random(10, 10)
    print(a)
    # # 写入题库信息
    # s = [["任务规划室",["德克萨斯雷达","的考试辅导"]],["任务规划室1",["德克萨斯雷达","的考试辅导"]]]
    # writejson("data/tkinfo.json", s)

    # tkinfo = readjson("data/tkinfo1.json")
    # print(tkinfo)
    # print(tkinfo.keys())

    # danwei = list(tkinfo.keys())
    # print(danwei)

    # s = [k[0] for k in p]
    #
    # print(s)


# class MyTool:
#     def readjson(self, str_filepath):
#         try:
#             with open(str_filepath, "r") as f:
#                 data = json.load(f)
#                 return data
#         except:
#             print("文件读取错误！")
#
#     def writejson(self, str_filename):
#         # try:
#         #     with open(str_filename, "w") as f:
#         #         data = json.dump()
#         #         return data
#         # except:
#         #     print("文件读取错误！")
#         pass
#
#     def animation(self):
#         pass